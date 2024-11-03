import os
from openpyxl import Workbook, load_workbook
from config import RPi


class Excel:
    def __init__(self, file=RPi.EXCEL_FILENAME):
        self.file = file
        if os.path.exists(self.file):
            self.wb = load_workbook(self.file)

        else:
            self.wb = Workbook()
            self.wb.active.title = "ENV"
            self.wb.create_sheet(title="LEAF")
            self.wb.create_sheet(title="SOIL")
            self.wb.save(self.file)
            print("New excel workbook created: {}".format(self.file))

    def add_row(self, sheet_name, row):
        sheet = self.wb[sheet_name]
        sheet.append(row)

    def save(self):
        self.wb.save(self.file)
        print("Data saved to file: {}".format(self.file))

